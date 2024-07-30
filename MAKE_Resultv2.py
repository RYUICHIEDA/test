import matplotlib.pyplot as plt
import os
import numpy as np
import io
from openpyxl import Workbook
from openpyxl.drawing.image import Image
from openpyxl.utils import get_column_letter



global folder_path
global Light_Lambda
global xpixel
global ypixel



# 光の波長をリストの番号に変える関数(中の数値はSpecimIQの取得する光波長の範囲と分解能に依存)
def Light_Lambda_to_Light_Lambda_number(Light_Lambda_value):
    return int(round((Light_Lambda_value - 397.32) / ((1003.58 - 397.32) / 203), 0))

# フォルダパスから指定した範囲のnpyファイルを読み込み、4次元データを作成
# 第一要素は走査数（光源の位置が右から左の順に格納）、第二要素は任意の光波長λ
# 第三要素はx軸のピクセル、格納されるデータは光強度   
def process_files(st_num, en_num):
    data = np.zeros((en_num - st_num + 1, len(Light_Lambda), xpixel, ypixel))
    
    for number, data_index in enumerate(range(st_num, en_num + 1)):
        file_path = os.path.join(folder_path, f"{data_index:03d}_processed.npy")
        if os.path.exists(file_path):
            l_data = np.load(file_path)
            for lam, Light_Lambda_value in enumerate(Light_Lambda):
                data[number, lam, :, :] = l_data[:, :, Light_Lambda_to_Light_Lambda_number(Light_Lambda_value)]
    
    data = np.flip(data, axis=2)
    
    return data
'''
def ave_eva(data, sp_xscan_num, sp_yscan_num, ave_range):
    eva_data = np.zeros((data.shape[0], data.shape[1], data.shape[3]))
    for i in range(data.shape[0]):
        for j in range(data.shape[1]):
            for k in range(data.shape[3]):
                eva_data[i, j, k] = data[]
'''
# 輝度のカラーマップを描画するプログラム
def draw_grayscale(arr, file_num, xst, xen, yst, yen, cap):
    image_binary = []
    # 配列の要素の大きさに応じて色を変換する関数を定義
    cmap = plt.get_cmap('gray')
    for j in range(len(Light_Lambda)):
        arr_normalized = (arr[j,:,:] - np.min(arr[j,:,:])) / (np.max(arr[j,:,:]) - np.min(arr[j,:,:]))
        # グラフをプロット
        plt.imshow(arr_normalized.T, cmap=cmap ,extent=[xst, xen, yen, yst])

        # 軸ラベル（入れ替えたのでラベルも入れ替える）
        plt.title(f'{cap}グレースケール画像 \n (光波長[nm]　λ = {Light_Lambda[j]}) \n ファイル番号 : {file_num}', fontname ='MS Gothic') 
        plt.xlabel('x [pixel]')
        plt.ylabel('y [pixel]')

        plt.yticks(range(yst, yen+1, int((yen-yst)/10)))

        # カラーバー
        plt.colorbar(label='Light Intencity')

        # グラフの保存
        save_filename = f'{glaph_save_path}\\No{file_num}_λ{Light_Lambda[j]}_{cap}グレースケール画像.png'
        plt.savefig(save_filename, bbox_inches='tight')  # bbox_inches='tight'で余白を最小限にして保存

        # メモリ上に画像を保存
        image_buffer = io.BytesIO()
        plt.savefig(image_buffer, format='png', bbox_inches='tight')
        image_buffer.seek(0)

        # 画像をバイナリデータとして取得
        image_binary.append(image_buffer.read())
        image_buffer.close()

        plt.clf()

    return image_binary


def insert_images_to_excel(image_datas, excel_file, start_row, ws):
    # Excelファイルが存在しない場合は新規作成
    if not os.path.exists(excel_file):
        wb = Workbook()
        wb.save(excel_file)

    # 画像を挿入するセルの行と列を指定
    current_row = start_row
    current_column = 2

    for image_data in image_datas:
        # 画像データをExcelファイルに挿入
        img = Image(io.BytesIO(image_data))
        cell = ws.cell(row=current_row, column=current_column)
        ws.add_image(img, cell.coordinate)

        # 画像のサイズに合わせてセルの大きさを変更
        ws.row_dimensions[current_row].height = img.height
        ws.column_dimensions[get_column_letter(current_column)].width = img.width / 7.5  # 仮の調整係数

        # 次の画像を挿入する位置を更新
        current_column += 1
    
    current_row += 1

    return current_row





# 定数（書き換えないでください）
mpp = 0.15152 # 1ピクセルあたりの長さ[mm]　画像から判読して決定しているため、後に再調査して変更する必要あり

# 画像のxとyのピクセル数
xpixel = 512
ypixel = 512

row = 2
column = 2
wb = Workbook()
ws = wb.active

# 以下に囲っている部分がユーザーが変更できる箇所
#----------------------------------------------------------------------------------------------------------------------------------

# ファイルのパス
folder_path = r'C:\Users\edaryuuichi\Desktop\240322\D\output' # 000_processed.npyがあるフォルダのパスを指定してください
# 作成したグラフを保存するファイルパス
glaph_save_path = os.path.join(os.path.dirname(folder_path), 'glaph')
Excel_path = os.path.join(glaph_save_path, "output.xlsx")
if not os.path.exists(glaph_save_path): os.makedirs(glaph_save_path)

# 全光照射画像のファイル番号
ensu_num =  357

# スポット光照射画像のファイル番号
sp_st_num = 358     # 開始番号
sp_en_num = 375     # 終了番号

yst = 210           # yの領域の始点
yen = 250           # yの領域の終点

sp_xscan_num = 9
sp_yscan_num = int((sp_en_num - sp_st_num + 1) / sp_xscan_num)

x_center = 265      # xpixelの中心位置
LS = 3              # x_centerから左方向の距離[mm]
RS = 5              # x_centerから右方向の距離[mm]

# 光の波長λ[nm]を指定するリスト　順不同でも可
Light_Lambda = [550,640,700,800]

#----------------------------------------------------------------------------------------------------------------------------------

# NumPy配列にデータを読み込む
ensu_data = process_files(ensu_num, ensu_num)

# LSとRSで指定した距離[mm]分x軸を動かしたときのピクセルを計算
ensu_xst = int(x_center - LS / mpp)
ensu_xen = int(x_center + RS / mpp)
img = draw_grayscale(ensu_data[0, :, ensu_xst:ensu_xen, yst:yen], ensu_num, ensu_xst, ensu_xen, yst, yen, cap = '全光照射')
row = insert_images_to_excel(img,Excel_path,row,ws)

sp_data = process_files(sp_st_num, sp_en_num)

row_group = []
for i in range(sp_yscan_num):
    st_row = row
    for j in range(sp_xscan_num):
        sp_xst = int(x_center - LS / mpp - (RS - j) / mpp)
        sp_xen = int(x_center + RS / mpp - (RS - j) / mpp)
        sp_yst = int(yst - i / mpp)
        sp_yen = int(yen - i / mpp)
        img = draw_grayscale(sp_data[j + i * sp_xscan_num, :, sp_xst:sp_xen, sp_yst:sp_yen], sp_st_num + j + i * sp_xscan_num, sp_xst, sp_xen, sp_yst, sp_yen, cap = 'スポット光照射')
        row = insert_images_to_excel(img,Excel_path,row,ws)
    row_group.append([st_row, row - 1])
    row += 1

for row_no in row_group:
    ws.row_dimensions.group(*row_no, outline_level=1, hidden=True)

# 写真番号の位置を示す配列　カラーマップ描画時に使用
steplist = range(-1 * LS, RS + 1)



wb.save(Excel_path)