import matplotlib.pyplot as plt
import os
import numpy as np
import io
from openpyxl import Workbook
from openpyxl.drawing.image import Image
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment
from matplotlib.colors import LogNorm

# グローバル変数
global folder_path
global glaph_save_path
global Light_Lambda
global xpixel, ypixel
global mpp
global sp_xscan_num, sp_yscan_num
global y_center
global pix_sp_yrange
global yst, yen
global steplist


# -LSからRSの範囲までdeltaごとに区切ったリストを作成
def make_step(LS, RS, delta):
    lst = []
    lstlen = int((LS + RS) / delta) + 1
    for i in range(lstlen):
        lst.append(-LS + i * delta)
    return lst

# 光の波長をリストの番号に変える関数(中の数値はSpecimIQの取得する光波長の範囲と分解能に依存)
def Light_Lambda_to_Light_Lambda_number(Light_Lambda_value):
    return int(round((Light_Lambda_value - 397.32) / ((1003.58 - 397.32) / 203), 0))

# 指定したピクセル数の範囲内で左右に分割した時の差分の総和を返す関数
# 24/04/16現在、スポット光中心がx=265[pixel]だったため、処理を変更しています コメントアウトしている部分が本来の処理
# 24/05/14　スポット光の中心位置が測定装置の位置によって変化することを鑑みると、事前にスポット光中心位置x_centerを指定した方がよいのでないだろうか
def dif(data, x_center, cal_ran):
    dif = 0
    for i in range(int((cal_ran / 2) / mpp)):
        L_index = x_center - 1 - i
        R_index = x_center + i
        #L_index = len(data)//2 - 1 - i
        #R_index = len(data)//2 + i
        dif += abs(data[L_index] - data[R_index])
    return dif

# フォルダパスから指定した範囲のnpyファイルを読み込み、4次元データを作成
# 以下、返り値dataの要素
# 1: 写真番号インデックス
# 2: 光波長インデックス
# 3: 画像のxピクセル
# 4: 画像のyピクセル
def process_files(st_num, en_num):
    data = np.zeros((en_num - st_num + 1, len(Light_Lambda), xpixel, ypixel))
    
    for number, data_index in enumerate(range(st_num, en_num + 1)):
        file_path = os.path.join(folder_path, f"{data_index:03d}_processed.npy")
        if os.path.exists(file_path):
            l_data = np.load(file_path)
            for lam, Light_Lambda_value in enumerate(Light_Lambda):
                data[number, lam, :, :] = l_data[:, :, Light_Lambda_to_Light_Lambda_number(Light_Lambda_value)]
    
    data = np.flip(data, axis=2)
    
    return data

# 4次元データdataを指定したmethodで処理し、3次元の評価データを出力する関数
# 以下、返り値eva_dataの要素
# 1: 写真番号インデックス
# 2: 光波長インデックス
# 3: 画像のyピクセル
def process_data(data, x_center, cal_ran, method):
    eva_data = np.zeros((data.shape[0], data.shape[1], data.shape[3]))
    for i in range(data.shape[0]):
        for j in range(data.shape[1]):
            for k in range(data.shape[3]):
                if method == 0:
                    eva_data[i, j, k] = np.mean(data[i, j, x_center - int(cal_ran / mpp):x_center + int(cal_ran / mpp), k])
                elif method == 1:
                    eva_data[i, j, k] = data[i, j, x_center, k]
                elif method == 2:
                    eva_data[i, j, k] = dif(data[i, j, :, k], x_center, cal_ran)
    return eva_data

# 輝度のカラーマップを描画するプログラム
def draw_grayscale(arr, file_num, xst, xen, yst, yen, cap):
    image_binary = []
    # 配列の要素の大きさに応じて色を変換する関数を定義
    cmap = plt.get_cmap('gray')
    for j in range(len(Light_Lambda)):
        #arr_normalized = (arr[j,:,:] - np.min(arr[j,:,:])) / (np.max(arr[j,:,:]) - np.min(arr[j,:,:]))
        # グラフをプロット
        plt.imshow(arr[j,:,:].T, aspect = 'auto', cmap = cmap, extent = [xst, xen, yen, yst])

        # 軸ラベル（入れ替えたのでラベルも入れ替える）
        plt.title(f'{cap}グレースケール画像 \n (光波長[nm]　λ = {Light_Lambda[j]}) \n ファイル番号 : {file_num}', fontname ='MS Gothic') 
        plt.xlabel('x [pixel]')
        plt.ylabel('y [pixel]')

        plt.yticks(range(yst, yen + 1, int((yen - yst) / 10)))

        # カラーバー
        plt.colorbar(label='Light Intencity')

        #plt.show()

        # グラフの保存
        save_filename = f'{glaph_save_path}\\No{file_num}_λ{Light_Lambda[j]}_{cap}グレースケール画像.png'
        plt.savefig(save_filename, bbox_inches='tight')  # bbox_inches='tight'で余白を最小限にして保存

        # メモリ上に画像を保存
        image_buffer = io.BytesIO()
        plt.savefig(image_buffer, format='png', bbox_inches='tight')
        image_buffer.seek(0)

        # 画像をバイナリデータとして取得
        image_binary.append(image_buffer.read())
        image_buffer.close()

        plt.clf()

    return image_binary

# 輝度のカラーマップを描画するプログラム
def draw_colormap(arr, stlist, yst, yen, st_num, sp_xscan_num, cap):
    image_binary = []
    # 配列の要素の大きさに応じて色を変換する関数を定義
    arr = np.flip(arr, axis = 2)
    arr = np.flip(arr, axis = 0)
    cmap = plt.get_cmap('Reds_r')

    for i in range(arr.shape[1]):
        # グラフをプロット
        plt.imshow(arr[:,i,:].T, cmap=cmap, origin='lower', interpolation='nearest',aspect='auto', extent=[stlist[0] - deltax / 2, stlist[-1] + deltax / 2, yen, yst])

        # 軸ラベル（入れ替えたのでラベルも入れ替える）
        plt.title(f'{cap}カラーマップ\n (光波長[nm]　λ = {Light_Lambda[i]}) \n ファイル番号 : {st_num} - {st_num + sp_xscan_num - 1}', fontname ='MS Gothic') 
        plt.xlabel('スポット光位置 [mm]', fontname ='MS Gothic')
        plt.ylabel('y [pixel]')

        plt.yticks(range(yst, yen + 1, int((yen - yst) / 10)))

        # カラーバー
        plt.colorbar(label = 'Light Intencity')

        # グラフの保存
        save_filename = f'{glaph_save_path}\\No{st_num}-{st_num + sp_xscan_num - 1}_λ{Light_Lambda[i]}_{cap}カラーマップ.png'
        plt.savefig(save_filename, bbox_inches='tight')  # bbox_inches='tight'で余白を最小限にして保存

        # メモリ上に画像を保存
        image_buffer = io.BytesIO()
        plt.savefig(image_buffer, format='png', bbox_inches='tight')
        image_buffer.seek(0)

        # 画像をバイナリデータとして取得
        image_binary.append(image_buffer.read())
        image_buffer.close()

        plt.clf()
    
    return image_binary

def normalize(arr):
    min_val = np.min(arr)
    max_val = np.max(arr)
    
    # すべての要素が同じ値の場合、正規化が不可能なのでそのまま返す
    if min_val == max_val:
        return arr
    
    normalized_array = (arr - min_val) / (max_val - min_val)
    return normalized_array

# 評価値によって作成したカラーマップをExcel上に並べるプログラム
def lineup_colormap(arr, row, cap):
    for i in range(sp_yscan_num):
        cell = ws.cell(row = row, column = 1, value = f"y = {0 - i}")
        cell.alignment = Alignment(vertical = 'top')
        sp_yst = yst + int((UPS - i) / mpp)
        sp_yen = yen + int((UPS - i) / mpp)
        img = draw_colormap(arr[i * sp_xscan_num : sp_xscan_num + i * sp_xscan_num, :, sp_yst:sp_yen], steplist, sp_yst, sp_yen, sp_st_num + i * sp_xscan_num, sp_xscan_num, cap = f'{cap}')
        row = insert_images_to_excel(img, Excel_path, row ,ws)
        row += 1
    return row

# 撮影した画像を縦に連結させてカラーマップ出力する
def y_merge_colormap(arr, row, cap):
    mer_data = np.zeros((sp_xscan_num, arr.shape[1], pix_sp_yrange * sp_yscan_num))
    for i in range(sp_yscan_num):
        ev_el = np.zeros((sp_xscan_num, arr.shape[1], pix_sp_yrange))
        if pix_sp_yrange % 2 != 1:
            for j in range(arr.shape[1]):
                mer_data[:, j, i * pix_sp_yrange:pix_sp_yrange + i * pix_sp_yrange] = normalize(arr[i * sp_xscan_num : sp_xscan_num + i * sp_xscan_num, j, y_center - int((pix_sp_yrange - 1) / 2):y_center + int((pix_sp_yrange - 1) / 2)])
        else:
            for j in range(pix_sp_yrange):
                ev_el[:, :, j] = (arr[i * sp_xscan_num : sp_xscan_num + i * sp_xscan_num, :, y_center - int(pix_sp_yrange / 2) + j] + arr[i * sp_xscan_num : sp_xscan_num + i * sp_xscan_num, :, y_center - int(pix_sp_yrange / 2) + i + 1]) / 2
            for j in range(arr.shape[1]):
                mer_data[:, j, i * pix_sp_yrange:pix_sp_yrange + i * pix_sp_yrange] = normalize(ev_el[:, j, :])            
    img = draw_colormap(mer_data[:, :, :], steplist, 0, mer_data.shape[2], 1, 0, cap = f'{cap}')
    row = insert_images_to_excel(img, Excel_path, row ,ws)
    row += 1
    return row



# Excelファイルに作成したグラフを貼り付ける
def insert_images_to_excel(image_datas, excel_file, start_row, ws):
    # Excelファイルが存在しない場合は新規作成
    if not os.path.exists(excel_file):
        wb = Workbook()
        wb.save(excel_file)

    # 画像を挿入するセルの行と列を指定
    current_row = start_row
    current_column = 2

    for image_data in image_datas:
        # 画像データをExcelファイルに挿入
        img = Image(io.BytesIO(image_data))
        cell = ws.cell(row=current_row, column=current_column)
        ws.add_image(img, cell.coordinate)

        # 画像のサイズに合わせてセルの大きさを変更
        ws.row_dimensions[current_row].height = img.height
        ws.column_dimensions[get_column_letter(current_column)].width = img.width / 7.5  # 仮の調整係数

        # 次の画像を挿入する位置を更新
        current_column += 1
    
    current_row += 1

    return current_row





# 定数（書き換えないでください）
mpp = 0.15152 # 1ピクセルあたりの長さ[mm]　画像から判読して決定しているため、後に再調査して変更する必要あり

# 画像のxとyのピクセル数
xpixel = 512
ypixel = 512

row = 2         # Excelの行の初期位置
column = 2      # Excelの列の位置（未実装）
wb = Workbook()
ws = wb.active

# 以下に囲っている部分がユーザーが変更できる箇所
#----------------------------------------------------------------------------------------------------------------------------------

# ファイルのパス
folder_path = r"C:\Users\edaryuuichi\Desktop\240605\output" # 000_processed.npyがあるフォルダのパスを指定してください
# 作成したグラフを保存するファイルパス
glaph_save_path = os.path.join(os.path.dirname(folder_path), 'glaph')
Excel_path = os.path.join(glaph_save_path, "output.xlsx")
if not os.path.exists(glaph_save_path): os.makedirs(glaph_save_path)

# 全光照射画像のファイル番号
ensu_num =  513

# スポット光照射画像のファイル番号
sp_st_num = 514     # 開始番号
sp_en_num = 530     # 終了番号

# 撮影条件

sp_xscan_num = 17   # x方向の撮影枚数
sp_yscan_num = 1    # y方向の撮影枚数

deltax = 0.5        # xのステップ区切り[mm]
deltay = 1          # yのステップ区切り[mm]（未実装）

UPS = 0             # y_centerから上方向の距離[mm]
UNS = 0             # y_centerから下方向の距離[mm]

LS = 4              # x_centerから左方向の距離[mm]
RS = 4              # x_centerから右方向の距離[mm]

yst = 210           # yの領域の始点
yen = 250           # yの領域の終点

# 描画範囲など

x_center = 262      # xpixelの中心位置
y_center = 232      # ypixelの中心位置

sp_yrange = 1.5     # y軸方向にカラーマップを重ね合わせる時の取得する範囲[mm]

# 光の波長λ[nm]を指定するリスト　順不同でも可
Light_Lambda = [400,450,500,550,600,650,700,750,800,850,900,950,1000]

#----------------------------------------------------------------------------------------------------------------------------------
pix_sp_yrange = int(sp_yrange/mpp)  # sp_yrangeをピクセル単位に変換した値
'''
# NumPy配列に全光照射画像データを読み込む　ensu:ENtire SUrface
ensu_data = process_files(ensu_num, ensu_num)

# LSとRSで指定した距離[mm]分x軸を動かしたときのピクセルを計算
ensu_xst = x_center - int(LS / mpp)
ensu_xen = x_center + int(RS / mpp)
ensu_yst = yst - int(UPS / mpp)
ensu_yen = yen + int(UNS / mpp)

# 全光照射画像を描画してExcel上に保存
img = draw_grayscale(ensu_data[0, :, ensu_xst:ensu_xen, ensu_yst:ensu_yen], ensu_num, xst = ensu_xst, xen = ensu_xen, yst = ensu_yst, yen = ensu_yen, cap = '全光照射')
row = insert_images_to_excel(img,Excel_path,row,ws)
print("全光照射描画完了")

# スポット光照射のデータをsp_dataに格納
sp_data = process_files(sp_st_num, sp_en_num)

row_group = [] # Excel上で画像群をグループ化するためのリスト

# 血栓を中心と仮定して、スポット光照射画像のグレースケールデータを光波長ごとに出力する
for i in range(sp_yscan_num):
    cell = ws.cell(row = row, column = 1, value = f"y = {0 - i}")
    cell.alignment = Alignment(vertical = 'top')
    st_row = row
    for j in range(sp_xscan_num):
        sp_xst = int(x_center - LS / mpp - (RS - j * deltax) / mpp)
        sp_xen = int(x_center + RS / mpp - (RS - j * deltax) / mpp)
        sp_yst = int(yst - i / mpp + UPS / mpp)
        sp_yen = int(yen - i / mpp + UPS / mpp)
        img = draw_grayscale(sp_data[j + i * sp_xscan_num, :, sp_xst:sp_xen, sp_yst:sp_yen], sp_st_num + j + i * sp_xscan_num, xst = sp_xst, xen = sp_xen, yst = sp_yst, yen = sp_yen, cap = 'スポット光照射')
        row = insert_images_to_excel(img, Excel_path, row, ws)
    row_group.append([st_row, row - 1])
    row += 1

for row_no in row_group:
    ws.row_dimensions.group(*row_no, outline_level=1, hidden=True)

print("スポット光画像描画完了")


# 写真番号の位置を示す配列　カラーマップ描画時に使用
steplist = make_step(LS, RS, deltax)

# 輝度平均値のカラーマップ描画

ave_sp_data = process_data(sp_data, x_center, 0.5, method = 0)  # 生データを評価用データに変換
row = lineup_colormap(ave_sp_data, row, cap = '輝度平均値')      # yごとの2次元カラーマップを描画し、Excelに貼り付ける
row = y_merge_colormap(ave_sp_data, row, cap = '輝度平均合体')   # 上記で作成したカラーマップをy方向に連結させて描画する

# 輝度最大値のカラーマップ描画

max_sp_data = process_data(sp_data, x_center, 0.5, method = 1)  # 生データを評価用データに変換
row = lineup_colormap(max_sp_data, row, cap = '輝度最大値')      # yごとの2次元カラーマップを描画し、Excelに貼り付ける
row = y_merge_colormap(max_sp_data, row, cap = '輝度最大合体')   # 上記で作成したカラーマップをy方向に連結させて描画する

# 差分カラーマップの描画

dif_sp_data = process_data(sp_data, x_center, 0.5, method = 2)  # 生データを評価用データに変換
row = lineup_colormap(dif_sp_data, row, cap = '差分')            # yごとの2次元カラーマップを描画し、Excelに貼り付ける
row = y_merge_colormap(dif_sp_data, row, cap = '差分合体')       # 上記で作成したカラーマップをy方向に連結させて描画する

print("評価画像描画完了")
'''
dum_dis = 1
renum_list = []
renum = (sp_xscan_num + 1) / 2

# 下側の数値を追加
lower_side = renum
while lower_side >= 1:
    renum_list.insert(0, lower_side - 1)
    lower_side -= dum_dis / deltax

# 上側の数値を追加
upper_side = renum
while upper_side <= sp_xscan_num:
    renum_list.append(upper_side - 1)
    upper_side += dum_dis / deltax

print(renum_list)

dum_data = np.zeros((len(renum_list, sp_data.shape[1], sp_data.shape[2], sp_data.shape[3])))

#wb.save(Excel_path)