import os
import matplotlib.pyplot as plt
import numpy as np
from openpyxl import Workbook
from openpyxl.drawing.image import Image
from openpyxl.utils import get_column_letter
import io

def create_image():
    # 仮のデータで画像を作成する例
    x = np.linspace(0, 10, 100)
    y = np.sin(x)

    plt.plot(x, y)
    plt.xlabel('X')
    plt.ylabel('Y')
    plt.title('Sample Plot')

    # メモリ上に画像を保存
    image_buffer = io.BytesIO()
    plt.savefig(image_buffer, format='png')
    image_buffer.seek(0)

    # 画像をバイナリデータとして取得
    image_binary = image_buffer.read()
    image_buffer.close()

    return image_binary

def insert_image_to_excel(image_data, excel_file, sheet_index, row, column):
    # Excelファイルが存在しない場合は新規作成
    if not os.path.exists(excel_file):
        wb = Workbook()
        wb.save(excel_file)

    # Excelファイルを開く
    wb = Workbook()
    ws = wb.worksheets[sheet_index]

    # 画像を挿入するセルの行と列を指定
    cell = ws.cell(row=row, column=column)

    # 画像データをExcelファイルに挿入
    img = Image(io.BytesIO(image_data))
    ws.add_image(img, cell.coordinate)

    # 画像のサイズに合わせてセルの大きさを変更
    ws.row_dimensions[row].height = img.height
    ws.column_dimensions[get_column_letter(column)].width = img.width / 7.5  # 仮の調整係数

    # Excelファイルを保存
    wb.save(excel_file)

# 画像を作成
image_data = create_image()

# Excelファイルのパス
excel_file = "output.xlsx"

# 挿入するシートのインデックス（0から始まる）
sheet_index = 0

# 画像を挿入するセルの行と列
row = 1
column = 1

# 画像をExcelファイルに挿入
insert_image_to_excel(image_data, excel_file, sheet_index, row, column)
